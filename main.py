import openpyxl

# Load the original Excel file
wb = openpyxl.load_workbook("C:\\Users\\optr6\\OneDrive\\Desktop\\junjuly2023.xlsx")
ws = wb["jun july 2023"]

selected_columns = [5, 3, 1]
start_row = 2
end_row = 5383

employee_data = {}
for row_index in range(start_row, end_row):
    row_data = []
    for col_index in selected_columns:
        cell_value = ws.cell(row=row_index, column=col_index).value
        row_data.append(cell_value)

    emp_id = row_data[0]
    entry_type = row_data[1]
    entry_time = row_data[2]

    if emp_id not in employee_data:
        employee_data[emp_id] = {"Employee": emp_id, "Entries": {"Entry-In": [], "Exit-Out": []}}

    if entry_type == "Entry-In":
        employee_data[emp_id]["Entries"]["Entry-In"].append(entry_time)
    elif entry_type == "Exit-Out":
        employee_data[emp_id]["Entries"]["Exit-Out"].append(entry_time)

new_wb = openpyxl.Workbook()
new_ws = new_wb.active

new_ws.append(["Employee", "Entry-In", "Exit-Out"])

for emp_id, data in employee_data.items():
    entry_in_list = data["Entries"]["Entry-In"]
    entry_out_list = data["Entries"]["Exit-Out"]
    max_entries = max(len(entry_in_list), len(entry_out_list))

    for i in range(max_entries):
        entry_in = entry_in_list[i] if i < len(entry_in_list) else ""
        entry_out = entry_out_list[i] if i < len(entry_out_list) else ""
        
        new_ws.append([data["Employee"], entry_in, entry_out])

# Save the new Excel file
new_wb.save("C:\\Users\\optr6\\OneDrive\\Desktop\\employee_data.xlsx")

# Close the original Excel file
wb.close()
